import pyodbc
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# === Connection: Azure SQL with AAD Password Authentication ===
conn = pyodbc.connect(
    "Driver={ODBC Driver 17 for SQL Server};"
    "Server=we-prd-operations-kpis.database.windows.net;"
    "Database=tss-subscriber-prd;"
    "UID=guanglin.si@marlink.com;"       
    "PWD=*2Aera5*cW5t-?Q;"                    
    "Authentication=ActiveDirectoryPassword;"
)

# === SQL Query: Top 5 delayed tickets assigned to RCS and unassigned ===
query = """
WITH Tickets AS (
    SELECT ProblemID
    FROM tblProblem
    WHERE AssignedToResourceID = 41
      AND AssignedToContactID IS NULL
      AND ProblemStatusID = 1
      AND CaseClosed = 0
      AND ResolvedBy IS NULL
),
LastAction AS (
    SELECT
        pa.ProblemID,
        MAX(pa.CreatedDateTime) AS LastActionTime
    FROM tblProblemAction pa
    INNER JOIN Tickets t ON pa.ProblemID = t.ProblemID
    GROUP BY pa.ProblemID
),
Delays AS (
    SELECT
        la.ProblemID,
        la.LastActionTime,
        DATEDIFF(MINUTE, la.LastActionTime, GETDATE()) AS IdleMinutes,
        GETDATE() AS SnapshotTime
    FROM LastAction la
)
SELECT TOP 5 *
FROM Delays
ORDER BY IdleMinutes DESC
"""

# === Run the query and read results into DataFrame ===
df = pd.read_sql(query, conn)
conn.close()

# === Excel File Setup ===
excel_file = r'C:\Users\gsi\Downloads\Dashboard\TicketswithDelayLog.xlsx'
sheet_name = 'Sheet1'  # ✅ change if your sheet is named differently

# === Append to Excel Table ===
book = load_workbook(excel_file)
sheet = book[sheet_name]
start_row = sheet.max_row + 1

with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False, header=False)

print(f"{datetime.now()} - ✅ Appended {len(df)} rows to Excel.")
